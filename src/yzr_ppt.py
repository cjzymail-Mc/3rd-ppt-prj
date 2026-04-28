#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""yzr_ppt.py — yzr 模板 PPT 评测页生成（零 pipeline 依赖）.

Self-contained: 所有工具函数都从 pipeline/03a + 03b + ppt_pipeline_common
拷贝/改写而来，运行时不依赖 pipeline/ 目录。

公开 API: make_codex_slide()
"""

from __future__ import annotations

import re
import sys
import time
import tempfile
import textwrap
from pathlib import Path
from typing import Any, Dict, List, Tuple

# 直接运行时（python src/yzr_ppt.py），将项目根目录加入 sys.path
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# src 内部依赖：GPT_5 + overlay helpers
GPT_5 = None
show_gpt_waiting_overlay = None
remove_gpt_waiting_overlay = None
try:
    from .Function_030 import GPT_5, show_gpt_waiting_overlay, remove_gpt_waiting_overlay  # type: ignore
except Exception:
    try:
        from src.Function_030 import GPT_5, show_gpt_waiting_overlay, remove_gpt_waiting_overlay  # type: ignore
    except Exception:
        try:
            from Function_030 import GPT_5, show_gpt_waiting_overlay, remove_gpt_waiting_overlay  # type: ignore
        except Exception:
            GPT_5 = None

# ---------------------------------------------------------------------------
# 共享纯数据工具（Fix2）：常量 / 评分 / 文本裁剪 / Excel 列工具
# 影响视觉输出的函数（_write_text / _write_chart / _apply_keyword_color /
# _build_rich_prompt 等）仍保留在本文件内，以保证 per-template 独立微调能力。
# ---------------------------------------------------------------------------
_shared_import_ok = False
try:
    from ._ppt_shared import (  # type: ignore
        _RED, _BLUE, _BLACK,
        _ADVANTAGE_MARKERS, _DISADVANTAGE_MARKERS,
        _find_col, _classify_columns, _col_values,
        _extract_score_means, _xlwings_to_rows,
        _score_10pt, _score_to_grade, _score_to_grade_letter,
        _score_to_grade_modifier, _sample_stat_text,
        clamp_text,
        _NAME_KEYWORDS, _WEIGHT_KEYWORDS,
        _com_get, _write_text, _write_chart,
        make_chart_for_yzr, _prepare_yzr_chart_data,
        _apply_keyword_color, _strip_bullet_on_section_headers,
    )
    _shared_import_ok = True
except Exception:
    pass
if not _shared_import_ok:
    try:
        from src._ppt_shared import (  # type: ignore
            _RED, _BLUE, _BLACK,
            _ADVANTAGE_MARKERS, _DISADVANTAGE_MARKERS,
            _find_col, _classify_columns, _col_values,
            _extract_score_means, _xlwings_to_rows,
            _score_10pt, _score_to_grade, _score_to_grade_letter,
            _score_to_grade_modifier, _sample_stat_text,
            clamp_text,
            _NAME_KEYWORDS, _WEIGHT_KEYWORDS,
            _com_get, _write_text, _write_chart,
            make_chart_for_yzr, _prepare_yzr_chart_data,
            _apply_keyword_color, _strip_bullet_on_section_headers,
        )
    except Exception:
        from _ppt_shared import (  # type: ignore
            _RED, _BLUE, _BLACK,
            _ADVANTAGE_MARKERS, _DISADVANTAGE_MARKERS,
            _find_col, _classify_columns, _col_values,
            _extract_score_means, _xlwings_to_rows,
            _score_10pt, _score_to_grade, _score_to_grade_letter,
            _score_to_grade_modifier, _sample_stat_text,
            clamp_text,
            _NAME_KEYWORDS, _WEIGHT_KEYWORDS,
            _com_get, _write_text, _write_chart,
            make_chart_for_yzr, _prepare_yzr_chart_data,
            _apply_keyword_color, _strip_bullet_on_section_headers,
        )

# Default GPT model
_MODEL = "openai/gpt-5.4"

# Clipboard copy-paste COM buffer (seconds)
_COPY_PASTE_DELAY = 1.5

# ---------------------------------------------------------------------------
# Hardcoded shape specs (yzr 标准模板在 Template 2.1.pptx 第 15 页)
# 注意：shape 名用 PPT 实际名称（英文前缀 Rectangle/Picture/TextBox），
# 仅 "图表 44" 保持中文（图表 shape 默认名）。
# ---------------------------------------------------------------------------
YZR_SHAPES = [
    {"name": "Rectangle 11", "strategy": "score_10pt"},
    {"name": "Rectangle 12", "strategy": "grade_letter_only"},
    {"name": "Rectangle 22", "strategy": "grade_modifier_only"},
    {"name": "Rectangle 17", "strategy": "sample_aggregation"},
    {"name": "Rectangle 19", "strategy": "skip"},
    {"name": "Picture 39",   "strategy": "extract_image"},
    {"name": "TextBox 16",   "strategy": "extract_column",
     "params": {"column": "鞋款名称"}},
    {"name": "Rectangle 68", "strategy": "gpt_prompted",
     "params": {"source": "补充说明", "filter": "缺点"},
     "budget": {"max_chars": 270, "max_lines": 9}},
    {"name": "Rectangle 77", "strategy": "gpt_prompted",
     "params": {"source": "补充说明", "filter": "优点"},
     "budget": {"max_chars": 201, "max_lines": 5}},
    {"name": "Chart 13",      "strategy": "mean_extraction"},
]

# yzr 标准模板所在页（Template 2.1.pptx 第 15 页）
_TEMPLATE_SLIDE = 15

# ---------------------------------------------------------------------------
# Tiny helper utilities (self-contained)
# ---------------------------------------------------------------------------

def _safe_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _numeric(v: Any):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _shoe_name(rows: List[List[Any]], col_hint: str = "") -> str:
    """Extract shoe name from Excel column."""
    keywords = ([col_hint] if col_hint else []) + ["鞋款名称", "鞋款", "试穿"]
    names = _col_values(rows, *keywords)
    unique = list(dict.fromkeys(_safe_text(n) for n in names if n))
    return unique[0] if unique else ""


# ---------------------------------------------------------------------------
# GPT prompt helpers (dynamic column version)
# ---------------------------------------------------------------------------

def _build_respondent_block(rows: List[List[Any]]) -> Tuple[str, int]:
    """Build a per-respondent data block for inclusion in GPT prompt.

    Dynamically matches columns via _classify_columns + _find_col — works
    with any Excel format (no hardcoded column names).
    Returns (block_text, respondent_count).
    """
    if not rows or len(rows) < 2:
        return "（无数据）", 0

    headers = [_safe_text(h) for h in rows[0]]
    n = len(rows) - 1

    name_col = _find_col(headers, _NAME_KEYWORDS)
    weight_col = _find_col(headers, _WEIGHT_KEYWORDS)
    score_cols, text_cols = _classify_columns(headers, rows)

    blocks = []
    for i, row in enumerate(rows[1:], 1):
        fd: Dict[str, str] = {}
        for j, h in enumerate(headers):
            if j < len(row):
                fd[h] = _safe_text(row[j])

        name = fd.get(name_col, f"受访者{i}") if name_col else f"受访者{i}"
        weight = fd.get(weight_col, "") if weight_col else ""

        scores = ", ".join(
            f"{h.split('（')[0] if '（' in h else h}={fd[h]}"
            for h in score_cols if fd.get(h)
        )
        feedbacks = "\n  ".join(
            f"{h.split('（')[0] if '（' in h else h}: {fd[h]}"
            for h in text_cols if fd.get(h)
        )

        weight_str = f"  体重:{weight}KG" if weight else ""
        parts = [f"【受访者{i}】{name}{weight_str}"]
        if scores:
            parts.append(f"  各项评分: {scores}")
        if feedbacks:
            parts.append(f"  试穿反馈:\n  {feedbacks}")
        blocks.append("\n".join(parts))

    return "\n\n".join(blocks), n


# prompt_src:  pipeline/prompt_templates/gpt_summary.md
# synced_at:   2026-04-16
# synced_by:   Developer（fix2 整改时从 pipeline 拷贝的最新版本）
# 定制说明：   yzr 模板无 p1p2 模式，style_anchor 始终传空；如需跟进 pipeline
#             最新改动，请手动 diff gpt_summary.md 并更新 synced_at。
def _build_rich_prompt(
    budget: dict,
    rows: List[List[Any]],
    focus: str = "",
    content_source: str = "补充说明",
    style_anchor: str = "",
) -> str:
    """Build GPT prompt for gpt_prompted shapes.

    focus: '优点' or '缺点' → free-form summarization mode.
    """
    respondent_block, n = _build_respondent_block(rows)
    max_chars = budget.get("max_chars", 200)
    max_lines = budget.get("max_lines", 6)
    extra = "每个分类不超过3行"

    if focus:
        task_line = (
            f"请从{n}名测试者的实际反馈中，自由归纳这款篮球鞋的【{focus}】。\n"
            f"根据实际反馈内容自行决定分段维度，不要按固定性能类别（如包裹性、止滑性等）分类。\n"
            f"在每条结论后注明（X/{n}）表示几分之几的测试者有此体验。\n"
            f"每段结论中，请将最核心的1-2个关键性能词用【】括起来（仅括词本身，不含标点），"
            f"例如：【止滑性】在室内场地表现稳定（2/{n}）。这些关键词后续会自动高亮显示。"
        )
        format_note = "- 参考文本仅作语调参考，不必复制其分类结构\n"
    else:
        task_line = (
            f"下面是{n}名测试者对这款篮球鞋的原始试穿反馈，"
            f"请帮我按分类汇总其中的【{content_source}】。\n"
            f"在每条结论后注明（X/{n}）表示几分之几的测试者有此反馈。\n"
            f"每段结论中，请将最核心的1-2个关键性能词用【】括起来（仅括词本身，不含标点），"
            f"例如：【止滑性】在室内场地表现稳定（2/{n}）。这些关键词后续会自动高亮显示。"
        )
        format_note = "- 严格按照参考文本的格式、语调、陈述方式\n"

    return (
        f"【参考文本（参考语调和信息密度）】\n{style_anchor}\n\n"
        f"【你的任务】\n{task_line}\n\n"
        f"注意：\n"
        f"- 你只能分析已有数据，不能推测或编造\n"
        f"- 直接给出结论，不要展示分析过程\n"
        f"- {format_note}"
        f"- 总字数控制在{max_chars}字左右\n"
        f"- 不超过{max_lines}行\n"
        f"- {extra}\n"
        f"- 结论中请自然融入：'样本'（如'本次{n}名样本'）、'反馈'（如'样本反馈'）、'建议'（末尾给出改进建议）\n\n"
        f"【{n}名测试者原始反馈】\n{respondent_block}\n\n"
        f"直接输出结论，不需要任何前言。"
    )


def _call_gpt(prompt: str, fallback: str, enabled: bool, model: str) -> str:
    """Call GPT_5 if enabled and available; return fallback otherwise."""
    if not enabled or GPT_5 is None:
        return fallback
    try:
        result = _safe_text(GPT_5(prompt, model))
        if result:
            return result
    except Exception:
        pass
    return fallback


# ---------------------------------------------------------------------------
# Image extraction (reads Excel via openpyxl, no COM conflict with xlwings)
# ---------------------------------------------------------------------------

def _extract_shoe_image(mc_sht) -> str:
    """Extract the first embedded image from the Excel questionnaire sheet."""
    try:
        from openpyxl import load_workbook
        excel_path = mc_sht.book.fullname
        wb = load_workbook(excel_path)
        for sname in wb.sheetnames:
            if "问卷" in sname:
                ws = wb[sname]
                imgs = getattr(ws, "_images", [])
                if imgs:
                    img = imgs[0]
                    fmt = getattr(img, "format", "png").lower() or "png"
                    tmp = Path(tempfile.mktemp(suffix=f".{fmt}"))
                    tmp.write_bytes(img._data())
                    return str(tmp)
    except Exception:
        pass
    return ""


# ---------------------------------------------------------------------------
# Content builder — routes by strategy
# ---------------------------------------------------------------------------

def _build_content(spec: dict, rows: List[List[Any]],
                   gpt_enabled: bool, model: str) -> str:
    """Build the text/data content for one shape spec."""
    strategy = spec["strategy"]
    params = spec.get("params", {})
    budget = spec.get("budget", {"max_chars": 80, "max_lines": 4})

    bad_txt = textwrap.dedent("""\
        【包裹性】
        前掌内/外侧弯折处均出现卡脚情况（3/3）
        鞋带孔生涩，在绑缚过程中较难对鞋带进行高效的收紧与放松（3/3）
        在进行内收步、剪刀步时，前掌外侧橡胶出现明显的挤压第五跖趾关节现象（1/3）
        【稳定性】
        在进行一些特定的动作时（行进间急停摆脱防守），出现轻微足外翻，具体位置在相比较前掌而言，后跟内侧支撑略低，但尚能接受（1/3）
        【止滑性】
        鞋底容易吸灰，导致打滑（3/3）
        鞋底在与地面摩擦时无[吱吱吱~]声响，对篮球爱好者来说安全感低。""")
    good_txt = textwrap.dedent("""\
        【止滑性】
        在室内干净木地板、室外水泥地上均能牢牢地锁定在地面。但在灰尘较大的木地板或塑胶场地上，需要频繁的擦拭鞋底灰尘。
        【场地感】
        静态下，前掌填覆材料异物感不明显。在动态下，前掌有明显的马蹄形[气垫]回弹反馈，在无碳板的结构下能给到足够澎湃的触地反馈。
        注：三种不同硬度的填覆材料，测试结果为：30C硬度下，所给到的脚感反馈最佳！""")

    if strategy == "skip":
        return ""

    if strategy == "score_10pt":
        score = _score_10pt(rows)
        if score is not None:
            return f"{score}/10"
        return "-.--/10"

    if strategy == "grade_letter":
        score = _score_10pt(rows)
        if score is not None:
            return _score_to_grade(score)
        return "B+"

    if strategy == "grade_letter_only":
        score = _score_10pt(rows)
        if score is not None:
            return _score_to_grade_letter(score)
        return "B"

    if strategy == "grade_modifier_only":
        score = _score_10pt(rows)
        if score is not None:
            return _score_to_grade_modifier(score)
        return "+"

    if strategy == "sample_aggregation":
        text = _sample_stat_text(rows)
        if not text:
            count = max(0, len(rows) - 1)
            text = f"试穿人数：{count}人"
        return text

    if strategy == "extract_column":
        col = params.get("column", "")
        name = _shoe_name(rows, col_hint=col)
        return name or "未知鞋款"

    if strategy == "gpt_prompted":
        focus = params.get("filter", "")
        src = params.get("source", "补充说明")
        fallback_map = {"优点": good_txt,
                        "缺点": bad_txt}
        fallback = fallback_map.get(focus, "样本反馈总体稳定，核心指标表现均衡。")
        prompt = _build_rich_prompt(budget, rows, focus=focus,
                                    content_source=src, style_anchor="")
        result = _call_gpt(prompt, fallback, gpt_enabled, model)
        # 安全网：GPT 返回后强制 clamp 到 budget 范围内
        max_chars = budget.get("max_chars", 200)
        max_lines = budget.get("max_lines", 6)
        return clamp_text(result, max_chars, max_lines)

    if strategy == "mean_extraction":
        means = _extract_score_means(rows)
        if not means:
            print("  [均值] _extract_score_means 返回空列表，使用占位数据")
            return "减震:0\n回弹:0\n稳定:0"
        print(f"  [均值] 提取到 {len(means)} 个指标均值: {[(k, round(v,2)) for k,v in means[:8]]}")
        return "\n".join(f"{k}:{v:.2f}" for k, v in means[:8])

    if strategy == "extract_image":
        return ""  # handled separately in make_codex_slide

    return ""


# ---------------------------------------------------------------------------
# COM write helpers: _com_get / _write_text / _write_chart
# 已移入 _ppt_shared.py，此处通过顶部 import 引入，不再本地定义。
# ---------------------------------------------------------------------------


# _apply_keyword_color 已移入 _ppt_shared.py（plan4），通过顶部 import 引入。


def _replace_image(slide, shp, img_path: str) -> None:
    """Replace an image shape with a picture from img_path (preserve aspect ratio)."""
    try:
        slot_left   = float(_com_get(shp, "Left",   0))
        slot_top    = float(_com_get(shp, "Top",    0))
        slot_width  = float(_com_get(shp, "Width",  100))
        slot_height = float(_com_get(shp, "Height", 100))
        name = _safe_text(_com_get(shp, "Name", ""))
        shp.Delete()

        new_shp = slide.Shapes.AddPicture(
            FileName=img_path,
            LinkToFile=False,
            SaveWithDocument=True,
            Left=slot_left, Top=slot_top,
            Width=-1, Height=-1,
        )

        nat_w = float(_com_get(new_shp, "Width",  slot_width))
        nat_h = float(_com_get(new_shp, "Height", slot_height))
        scale = min(slot_width / nat_w, slot_height / nat_h)
        new_w = nat_w * scale
        new_h = nat_h * scale

        new_shp.Width  = new_w
        new_shp.Height = new_h
        new_shp.Left   = slot_left  + (slot_width  - new_w) / 2
        new_shp.Top    = slot_top   + (slot_height - new_h) / 2

        try:
            new_shp.Name = name
        except Exception:
            pass
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def make_codex_slide(mc_sht, mc_ppt, mc_slide, sample_name: str,
                     mc_gpt: str = "n", mc_model: str = _MODEL):
    """Generate yzr evaluation slide, integrated into Main.py Copy->Paste flow.

    Clones Slide(15) (yzr 标准模板页) to end of presentation,
    then writes all shape content per YZR_SHAPES specs.

    Returns the new slide object (caller should update mc_slide).
    """
    gpt_enabled = (mc_gpt == "y")
    print(f"\n[yzr] 开始生成评测页  sample={sample_name}  gpt={'开启' if gpt_enabled else '关闭'}")
    rows = _xlwings_to_rows(mc_sht)
    print(f"[yzr] 读取问卷数据：{len(rows)} 行（含标题行），{len(rows[0]) if rows else 0} 列")

    # === Clone pattern — identical to all other sections in main.py ===
    X = mc_ppt.Slides.Count + 1
    print(f"[yzr] 克隆模板第 {_TEMPLATE_SLIDE} 页 → 新建第 {X} 页...")
    mc_ppt.Slides(_TEMPLATE_SLIDE).Copy()
    time.sleep(_COPY_PASTE_DELAY)
    new_slide = mc_ppt.Slides.Paste(X)
    time.sleep(1.0)

    # 显示等待遮罩（GPT 调用期间可见）
    _overlay = None
    if show_gpt_waiting_overlay is not None:
        try:
            _overlay = show_gpt_waiting_overlay(new_slide)
        except Exception:
            pass

    # Shape 位置微调 #fine_tuned
    try:
        _r68 = new_slide.Shapes("Rectangle 68")
        _r68.Left   = 24
        _r68.Top    = 264
        _r68.Width  = 415
        _r68.Height = 250
    except Exception:
        pass
    try:
        _r77 = new_slide.Shapes("Rectangle 77")
        _r77.Left   = 454
        _r77.Top    = 264
        _r77.Width  = 272
        _r77.Height = 226
    except Exception:
        pass
    # 鞋款名称 TextBox（用户实测坐标，2026-04-27）
    try:
        _tb16 = new_slide.Shapes("TextBox 16")
        _tb16.Left   = 20
        _tb16.Top    = 20
        _tb16.Width  = 200
        _tb16.Height = 36.35
    except Exception:
        pass

    # 确保 Rectangle 22 (评级 +/- 修饰符) 存在（2026-04-27）
    # Template 2.1.pptx 第 15 页中并无此 shape，YZR_SHAPES 循环找不到会跳过 →
    # PPT 上只显示 A、不显示 +/-。在此处自动创建 TextBox 并命名为 Rectangle 22，
    # 位置贴在 Rectangle 12（A）的右上角邻近，让循环后续能写入并写出 +/- 字符。
    try:
        new_slide.Shapes("Rectangle 22")  # 若已存在则跳过创建
    except Exception:
        try:
            _r12 = new_slide.Shapes("Rectangle 12")
            _r12_left  = float(_com_get(_r12, "Left",  0))
            _r12_top   = float(_com_get(_r12, "Top",   0))
            _r12_width = float(_com_get(_r12, "Width", 60))
            _mod_left   = _r12_left + _r12_width - 6   # 紧贴 A 右缘
            _mod_top    = _r12_top + 16                # 略低于 A 顶部，视觉居中偏上
            _mod_width  = 32
            _mod_height = 55
            # 第一个参数 1 = msoTextOrientationHorizontal
            _r22 = new_slide.Shapes.AddTextbox(1, _mod_left, _mod_top, _mod_width, _mod_height)
            _r22.Name = "Rectangle 22"
        except Exception as _e:
            print(f"  [警告] 自动创建 Rectangle 22 失败：{_e}")

    # === Per-shape content build and write ===
    print(f"[yzr] 开始逐 shape 写入，共 {len(YZR_SHAPES)} 个...")
    for spec in YZR_SHAPES:
        name     = spec["name"]
        strategy = spec["strategy"]

        if strategy == "skip":
            print(f"  [skip] {name}")
            continue

        # Find shape on the new slide
        shp = None
        try:
            shp = new_slide.Shapes(name)
        except Exception:
            pass
        if shp is None:
            print(f"  [未找到] {name}（模板中不存在此 shape，跳过）")
            continue

        print(f"  [处理] {name}  strategy={strategy}")

        # Special case: image extraction needs the sheet reference
        if strategy == "extract_image":
            img_path = _extract_shoe_image(mc_sht)
            if img_path:
                print(f"    图片路径: {img_path}")
                _replace_image(new_slide, shp, img_path)
            else:
                print(f"    未找到图片，跳过")
            continue

        # Build content
        content = _build_content(spec, rows, gpt_enabled, mc_model)

        # Route to correct writer
        if strategy == "mean_extraction" or bool(_com_get(shp, "HasChart", False)):
            # fix4: 分发场景 chart 走从零制表路线（xlwings 新建 + OLE 粘贴）
            # 不再用 _write_chart 原位改模板 chart 数据。
            # 路线决策与论据见 [feature03-transplant]/fix4（图表路线切换）.md
            try:
                L = float(_com_get(shp, "Left", 0))
                T = float(_com_get(shp, "Top", 0))
                W = float(_com_get(shp, "Width", 360))
                H = float(_com_get(shp, "Height", 170))
                shp.Delete()
            except Exception as _e:
                print(f"    [警告] 读取/删除模板 chart shape 失败: {_e}")
                continue

            # Chart 位置/尺寸微调 #fine_tuned
            # 用户实测调整后的最终坐标（2026-04-24），覆盖模板 Chart 13 原始位置
            # 读自 skills/read_selected_shape.py 输出（Chart 20 on slide 18）
            L = 242.19
            T = 21.95
            W = 467.24
            H = 224.48

            try:
                mc_cell = _prepare_yzr_chart_data(mc_sht, content)
            except Exception as _e:
                print(f"    [警告] 写入 Excel 临时数据失败: {_e}")
                continue

            try:
                _tmp_chart = make_chart_for_yzr(
                    mc_cell, new_slide, Left=L, Top=T, Width=W, Height=H,
                )
            except Exception as _e:
                print(f"    [警告] make_chart_for_yzr 失败: {_e}")
                continue

            # 清理 Excel：**只删 chart 对象，保留临时数据行**
            # 理由（用户实测经验）：
            #   - 即使 CutCopyMode = False 已执行，删除 Excel 端临时数据行仍会
            #     导致 PPT 端 OLE chart 数据丢失（bars 消失）
            #   - 保留临时数据虽然会让 Excel 稍显凌乱，但保证 PPT 产物稳定
            # 对照 Mc-debug-4.md line 744："我决定不折腾了，直接保留临时数据、保留图表吧。
            # 优先保证 ppt 图表的稳定性"
            _xl_app_api = None
            try:
                _xl_app_api = mc_sht.book.app.api
                _xl_app_api.DisplayAlerts = False
            except Exception:
                pass
            try:
                _tmp_chart.delete()
            except Exception:
                pass
            try:
                if _xl_app_api is not None:
                    _xl_app_api.DisplayAlerts = True
            except Exception:
                pass
        else:
            ok = _write_text(shp, content)
            if not ok:
                print(f"    [警告] _write_text 返回 False")
            if ok and strategy == "gpt_prompted":
                # section-aware coloring: 自动按上下文判断红/蓝/黑
                _apply_keyword_color(shp)

    # Rectangle 22 字体后处理：让 +/- 视觉上像"小一号的 A"
    # 模仿 Rectangle 12 (A) 的样式：微软雅黑 / 白色 / 加粗，仅字号显著小于 A
    # （A 的字号为 72，此处用 40 形成清晰的主次对比）
    try:
        _r22 = new_slide.Shapes("Rectangle 22")
        _tf = _r22.TextFrame
        _tf.AutoSize = 0
        _tr = _tf.TextRange
        _tr.Font.Name  = "微软雅黑"
        _tr.Font.Size  = 40
        _tr.Font.Bold  = True
        _tr.Font.Color.RGB = 0xFFFFFF  # 白色，与 A 一致
    except Exception:
        pass

    # 工作完成，删除等待遮罩
    if _overlay is not None and remove_gpt_waiting_overlay is not None:
        try:
            remove_gpt_waiting_overlay(_overlay)
        except Exception:
            pass

    print(f"[yzr] 完成！新页在第 {new_slide.SlideIndex} 页")
    return new_slide


# ---------------------------------------------------------------------------
# 单独调试入口：连接已打开的 Excel + PPT，只跑 yzr 这一页
# 用法：python src/yzr_ppt.py
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import os
    import win32com.client
    import xlwings

    # 项目根目录
    _proj_root = str(Path(__file__).resolve().parent.parent)

    # 打开 PPT 模板（与 Main.py 同方式，不自动保存）
    mc_app = win32com.client.Dispatch("PowerPoint.Application")
    mc_app.DisplayAlerts = 0
    mc_app.Visible = True
    mc_ppt = mc_app.Presentations.Open(_proj_root + r"\src\Template 2.1.pptx")
    mc_slide = mc_ppt.Slides(mc_ppt.Slides.Count)

    # 连接已打开的 Excel（问卷 sheet）
    mc_book = xlwings.books.active
    mc_sht = None
    for s in mc_book.sheets:
        if "问卷" in s.name:
            mc_sht = s
            break
    if mc_sht is None:
        print("未找到包含'问卷'的 sheet，请检查 Excel")
        exit(1)

    sample_name = mc_book.sheets["基础信息"].range("B2").value or "调试样品"

    print(f"[debug] sample: {sample_name}")
    print(f"[debug] sheet:  {mc_sht.name}")
    print(f"[debug] slides: {mc_ppt.Slides.Count}")

    mc_gpt = "n" #input("启用GPT? (y/n, 默认n): ").strip() or "n"

    new_slide = make_codex_slide(
        mc_sht, mc_ppt, mc_slide, sample_name,
        mc_gpt=mc_gpt, mc_model=_MODEL,
    )
    print(f"[debug] 完成！新页在第 {new_slide.SlideIndex} 页")
    print(f"[debug] 注意：模板文件未保存，请手动检查后关闭（不要保存）")
