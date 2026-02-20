#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Codex evaluation slide builder — zero pipeline/ dependency.

Self-contained: all helpers are copied/adapted from pipeline/03a and 03b.
Only public API: make_codex_slide().
"""

from __future__ import annotations

import re
import time
import tempfile
from pathlib import Path
from typing import Any, List, Tuple

# The only src-internal dependency: GPT_5 (relative import when used as package,
# absolute import when run standalone)
GPT_5 = None
try:
    from .Function_030 import GPT_5  # type: ignore
except Exception:
    try:
        from src.Function_030 import GPT_5  # type: ignore
    except Exception:
        GPT_5 = None

# ---------------------------------------------------------------------------
# Colors (must match main.py globals)
# ---------------------------------------------------------------------------
_RED  = 255        # red   = RGB(255, 0, 0)
_BLUE = 15773696   # light_blue = RGB(0, 176, 240)

# Default GPT model
_MODEL = "openai/gpt-5.2"

# Clipboard copy-paste COM buffer (seconds)
_COPY_PASTE_DELAY = 1.5

# ---------------------------------------------------------------------------
# Hardcoded shape specs (from 01-shape_detail.md annotations)
# ---------------------------------------------------------------------------
CODEX_SHAPES = [
    {"name": "矩形 11",   "strategy": "score_10pt",        "color_hint": ""},
    {"name": "矩形 12",   "strategy": "grade_letter",       "color_hint": ""},
    {"name": "矩形 17",   "strategy": "sample_aggregation", "color_hint": ""},
    {"name": "矩形 19",   "strategy": "skip",               "color_hint": ""},
    {"name": "图片 39",   "strategy": "extract_image",      "color_hint": ""},
    {"name": "文本框 16", "strategy": "extract_column",     "color_hint": "",
     "params": {"column": "鞋款名称"}},
    {"name": "矩形 68",   "strategy": "gpt_prompted",       "color_hint": "blue",
     "params": {"source": "补充说明", "filter": "缺点"},
     "budget": {"max_chars": 270, "max_lines": 9}},
    {"name": "矩形 77",   "strategy": "gpt_prompted",       "color_hint": "red",
     "params": {"source": "补充说明", "filter": "优点"},
     "budget": {"max_chars": 201, "max_lines": 5}},
    {"name": "图表 44",   "strategy": "mean_extraction",    "color_hint": ""},
]

# ---------------------------------------------------------------------------
# Tiny helper utilities (self-contained, no pipeline import)
# ---------------------------------------------------------------------------

def _safe_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _numeric(v: Any):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _com_get(obj, attr: str, default=None):
    """Safe getattr for COM objects (getattr raises on COM objects)."""
    try:
        return getattr(obj, attr)
    except Exception:
        return default


def _to_rows(value: Any) -> List[List[Any]]:
    """Convert xlwings used_range.value to List[List[Any]]."""
    if value is None:
        return []
    if isinstance(value, tuple):
        rows = []
        for row in value:
            if isinstance(row, tuple):
                rows.append(list(row))
            else:
                rows.append([row])
        return rows
    if isinstance(value, list):
        if value and isinstance(value[0], list):
            return value
        return [value]
    return [[value]]


def _extract_score_means(rows: List[List[Any]]) -> List[Tuple[str, float]]:
    """Extract per-column score means for bar charts.

    BUG FIX: score_like and backup_numeric are mutually exclusive (elif branches),
    so returning only score_like when non-empty drops all unmatched score columns.
    E.g. "缓震性", "包裹性", "抗扭转性", "防侧翻性", "耐久性" have no keyword match
    and land in backup_numeric — their means must be included in the overall average.
    Return all in-range columns: keyword-matched first, then unmatched.
    """
    if not rows or len(rows) < 2:
        return []

    headers = [_safe_text(h) for h in rows[0]]
    data = rows[1:]
    ncol = max(len(r) for r in rows)

    score_like = []
    backup_numeric = []
    score_keys = [
        "评分", "分数", "打分", "满意", "体验", "表现",
        "减震", "回弹", "稳定", "抓地", "舒适", "透气", "支撑",
    ]
    reject_keys = ["姓名", "昵称", "电话", "联系方式", "地址", "微信", "备注", "日期", "时间"]

    for c in range(ncol):
        header = headers[c] if c < len(headers) else f"指标{c+1}"
        if any(k in header for k in reject_keys):
            continue

        vals = []
        for r in data:
            if c >= len(r):
                continue
            n = _numeric(r[c])
            if n is not None:
                vals.append(float(n))

        if not vals:
            continue

        mean_val = sum(vals) / len(vals)
        in_score_range = all(0 <= v <= 20 for v in vals)
        if any(k in header for k in score_keys) and in_score_range:
            score_like.append((header, round(mean_val, 3)))
        elif in_score_range:
            backup_numeric.append((header, round(mean_val, 3)))

    return score_like + backup_numeric


# ---------------------------------------------------------------------------
# Data extraction helpers (ported from 03a_build_shape.py)
# ---------------------------------------------------------------------------

def _col_values(rows: List[List[Any]], *keywords: str) -> List[Any]:
    """Return all non-None values from the first column whose header
    contains any of the given keywords."""
    if not rows:
        return []
    headers = [_safe_text(h) for h in rows[0]]
    for kw in keywords:
        for idx, h in enumerate(headers):
            if kw in h:
                return [
                    row[idx]
                    for row in rows[1:]
                    if idx < len(row) and row[idx] is not None
                    and _safe_text(row[idx]) != ""
                ]
    return []


def _score_10pt(rows: List[List[Any]]):
    """Calculate overall mean score, auto-detects 5-scale vs 10-scale,
    returns a 10-point float or None."""
    means = _extract_score_means(rows)
    if not means:
        return None
    overall = sum(v for _, v in means) / len(means)
    max_mean = max(v for _, v in means)
    if max_mean <= 5.5:
        return round(overall * 2, 2)
    return round(overall, 2)


def _score_to_grade(score_10: float) -> str:
    """Convert a 10-point score to a letter grade."""
    s = score_10 * 10
    if s >= 95: return "S+"
    if s >= 90: return "S-"
    if s >= 85: return "A+"
    if s >= 80: return "A-"
    if s >= 75: return "B+"
    if s >= 70: return "B-"
    if s >= 65: return "C+"
    return "C-"


def _sample_stat_text(rows: List[List[Any]]) -> str:
    """Build sample stat text: trial count / avg weight / court position."""
    count = max(0, len(rows) - 1)

    weights = _col_values(rows, "体重", "Weight", "重量")
    valid_w = [float(w) for w in weights if _numeric(w) is not None]
    avg_w = round(sum(valid_w) / len(valid_w), 1) if valid_w else None

    positions = _col_values(rows, "球场定位", "打法", "定位", "位置")
    pos_clean: List[str] = []
    seen: set = set()
    for p in positions:
        s = _safe_text(p)
        if s and s not in seen:
            seen.add(s)
            pos_clean.append(s)

    lines = [f"试穿人数：{count}人"]
    if avg_w is not None:
        lines.append(f"测试者平均体重：{avg_w}KG")
    if pos_clean:
        lines.append(f"测试者球场定位：{'、'.join(pos_clean)}")
    return "\n".join(lines)


def _shoe_name(rows: List[List[Any]], col_hint: str = "") -> str:
    """Extract shoe name from Excel column."""
    keywords = ([col_hint] if col_hint else []) + ["鞋款名称", "鞋款", "试穿"]
    names = _col_values(rows, *keywords)
    unique = list(dict.fromkeys(_safe_text(n) for n in names if n))
    return unique[0] if unique else ""


# ---------------------------------------------------------------------------
# GPT prompt helpers (ported from 03a_build_shape.py)
# ---------------------------------------------------------------------------

_SCORE_COLS = [
    "抓地性（Traction）", "缓震性（Cushioning）", "包裹性（Lockdwon）",
    "抗扭转性（Torsional Support ）", "重量&透气性（Weight&Ventilation）",
    "防侧翻性（Lateral Stability）", "耐久性（Durability）",
]
_TEXT_COLS = [
    "你的脚型", "你认为这双鞋是否有足够的足弓支撑",
    "你认为这双鞋更加适合什么打法的球员穿", "补充说明",
]


def _build_respondent_block(rows: List[List[Any]]) -> Tuple[str, int]:
    """Build a per-respondent data block for inclusion in GPT prompt."""
    if not rows or len(rows) < 2:
        return "（无数据）", 0

    headers = [_safe_text(h) for h in rows[0]]
    n = len(rows) - 1
    blocks = []

    for i, row in enumerate(rows[1:], 1):
        fd: dict = {}
        for j, h in enumerate(headers):
            if j < len(row):
                fd[h] = _safe_text(row[j])

        name = fd.get("姓名（name）", f"受访者{i}")
        weight = fd.get("体重（Weight）", "")

        scores = ", ".join(
            f"{h.split('（')[0]}={fd[h]}"
            for h in _SCORE_COLS if fd.get(h)
        )
        feedbacks = "\n  ".join(
            f"{h.split('（')[0] if '（' in h else h}: {fd[h]}"
            for h in _TEXT_COLS if fd.get(h)
        )
        parts = [f"【受访者{i}】{name}  体重:{weight}KG"]
        if scores:
            parts.append(f"  各项评分: {scores}")
        if feedbacks:
            parts.append(f"  试穿反馈:\n  {feedbacks}")
        blocks.append("\n".join(parts))

    return "\n\n".join(blocks), n


def _build_rich_prompt(
    budget: dict,
    rows: List[List[Any]],
    focus: str = "",
    content_source: str = "补充说明",
    style_anchor: str = "",
) -> str:
    """Build GPT prompt for gpt_prompted shapes.

    focus: '优点' or '缺点' → free-form summarization mode.
    """
    respondent_block, n = _build_respondent_block(rows)
    max_chars = budget.get("max_chars", 200)
    max_lines = budget.get("max_lines", 6)
    extra = "每个分类不超过3行"

    if focus:
        task_line = (
            f"请从{n}名测试者的实际反馈中，自由归纳这款篮球鞋的【{focus}】。\n"
            f"根据实际反馈内容自行决定分段维度，不要按固定性能类别（如包裹性、止滑性等）分类。\n"
            f"在每条结论后注明（X/{n}）表示几分之几的测试者有此体验。\n"
            f"每段结论中，请将最核心的1-2个关键性能词用【】括起来（仅括词本身，不含标点），"
            f"例如：【止滑性】在室内场地表现稳定（2/{n}）。这些关键词后续会自动高亮显示。"
        )
        format_note = "- 参考文本仅作语调参考，不必复制其分类结构\n"
    else:
        task_line = (
            f"下面是{n}名测试者对这款篮球鞋的原始试穿反馈，"
            f"请帮我按分类汇总其中的【{content_source}】。\n"
            f"在每条结论后注明（X/{n}）表示几分之几的测试者有此反馈。"
        )
        format_note = "- 严格按照参考文本的格式、语调、陈述方式\n"

    return (
        f"【参考文本（参考语调和信息密度）】\n{style_anchor}\n\n"
        f"【你的任务】\n{task_line}\n\n"
        f"注意：\n"
        f"- 你只能分析已有数据，不能推测或编造\n"
        f"- 直接给出结论，不要展示分析过程\n"
        f"- {format_note}"
        f"- 总字数控制在{max_chars}字左右\n"
        f"- 不超过{max_lines}行\n"
        f"- {extra}\n"
        f"- 结论中请自然融入：'样本'（如'本次{n}名样本'）、'反馈'（如'样本反馈'）、'建议'（末尾给出改进建议）\n\n"
        f"【{n}名测试者原始反馈】\n{respondent_block}\n\n"
        f"直接输出结论，不需要任何前言。"
    )


def _call_gpt(prompt: str, fallback: str, enabled: bool, model: str) -> str:
    """Call GPT_5 if enabled and available; return fallback otherwise."""
    if not enabled or GPT_5 is None:
        return fallback
    try:
        result = _safe_text(GPT_5(prompt, model))
        if result:
            return result
    except Exception:
        pass
    return fallback


# ---------------------------------------------------------------------------
# xlwings → rows converter
# ---------------------------------------------------------------------------

def _xlwings_to_rows(mc_sht) -> List[List[Any]]:
    """xlwings Sheet → List[List[Any]], using CurrentRegion of the data anchor.

    Uses get_range() (from Function_030) to find the questionnaire anchor cell,
    then reads only its CurrentRegion — excludes temp/residual rows that are
    separated from the data by empty rows but still appear in used_range.
    Falls back to used_range if get_range is unavailable.
    """
    try:
        import importlib
        fn030 = importlib.import_module("src.Function_030")
        mc_cell0 = fn030.get_range(mc_sht)
        if mc_cell0 is not None:
            raw = mc_cell0.api.CurrentRegion.Value
            if raw is not None:
                return _to_rows(raw)
    except Exception:
        pass
    return _to_rows(mc_sht.used_range.value)


# ---------------------------------------------------------------------------
# Image extraction (reads Excel via openpyxl, no COM conflict with xlwings)
# ---------------------------------------------------------------------------

def _extract_shoe_image(mc_sht) -> str:
    """Extract the first embedded image from the Excel questionnaire sheet.

    Uses openpyxl read-only (does not conflict with xlwings COM instance).
    Returns the temp file path, or "" if not found.
    """
    try:
        from openpyxl import load_workbook
        excel_path = mc_sht.book.fullname   # xlwings -> full file path
        wb = load_workbook(excel_path)
        for sname in wb.sheetnames:
            if "问卷" in sname:
                ws = wb[sname]
                imgs = getattr(ws, "_images", [])
                if imgs:
                    img = imgs[0]
                    fmt = getattr(img, "format", "png").lower() or "png"
                    tmp = Path(tempfile.mktemp(suffix=f".{fmt}"))
                    tmp.write_bytes(img._data())
                    return str(tmp)
    except Exception:
        pass
    return ""


# ---------------------------------------------------------------------------
# Content builder — routes by strategy
# ---------------------------------------------------------------------------

def _build_content(spec: dict, rows: List[List[Any]],
                   gpt_enabled: bool, model: str) -> str:
    """Build the text/data content for one shape spec."""
    strategy = spec["strategy"]
    params = spec.get("params", {})
    budget = spec.get("budget", {"max_chars": 80, "max_lines": 4})

    if strategy == "skip":
        return ""

    if strategy == "score_10pt":
        score = _score_10pt(rows)
        if score is not None:
            return f"{score}/10"
        return "-.--/10"

    if strategy == "grade_letter":
        score = _score_10pt(rows)
        if score is not None:
            return _score_to_grade(score)
        return "B+"

    if strategy == "sample_aggregation":
        text = _sample_stat_text(rows)
        if not text:
            count = max(0, len(rows) - 1)
            text = f"试穿人数：{count}人"
        return text

    if strategy == "extract_column":
        col = params.get("column", "")
        name = _shoe_name(rows, col_hint=col)
        return name or "未知鞋款"

    if strategy == "gpt_prompted":
        focus = params.get("filter", "")
        src = params.get("source", "补充说明")
        fallback_map = {"优点": "样本反馈总体稳定，核心指标表现均衡。",
                        "缺点": "反馈集中，建议围绕关键指标继续优化。"}
        fallback = fallback_map.get(focus, "样本反馈总体稳定，核心指标表现均衡。")
        prompt = _build_rich_prompt(budget, rows, focus=focus,
                                    content_source=src, style_anchor="")
        return _call_gpt(prompt, fallback, gpt_enabled, model)

    if strategy == "mean_extraction":
        means = _extract_score_means(rows)
        if not means:
            return "减震:0\n回弹:0\n稳定:0"
        return "\n".join(f"{k}:{v:.2f}" for k, v in means[:8])

    if strategy == "extract_image":
        return ""  # handled separately in make_codex_slide

    return ""


# ---------------------------------------------------------------------------
# COM write helpers (ported from 03b_build_ppt_com.py)
# ---------------------------------------------------------------------------

def _write_text(shp, content: str) -> bool:
    """Write text to shape. Returns True on success."""
    if not bool(_com_get(shp, "HasTextFrame", 0)):
        return False
    tf = _com_get(shp, "TextFrame", None)
    tr = _com_get(tf, "TextRange", None) if tf is not None else None
    if tr is None:
        return False
    try:
        tf.AutoSize = 0  # ppAutoSizeNone — preserve template geometry
    except Exception:
        pass
    try:
        tr.Text = content
        return True
    except Exception:
        return False


def _write_chart(shp, content: str) -> bool:
    """Write chart data via SeriesCollection. Returns True on success.

    Strategy: BreakLink() in isolated try, then write Values+XValues directly.
    """
    chart = _com_get(shp, "Chart", None)
    if chart is None:
        return False

    lines = [x.strip() for x in (content or "").splitlines() if x.strip()]
    labels, values = [], []
    for line in lines[:10]:
        if ":" in line:
            k, v = line.rsplit(":", 1)
            labels.append(k.strip())
            try:
                values.append(float(v.strip()))
            except Exception:
                values.append(0.0)

    if not labels:
        return False

    try:
        try:
            chart.ChartData.Activate()
            time.sleep(0.5)
            chart.ChartData.BreakLink()
            time.sleep(0.3)
        except Exception:
            pass
        series = chart.SeriesCollection(1)
        series.Values = tuple(values)
        series.XValues = tuple(labels)
        return True
    except Exception:
        return False


def _apply_keyword_color(shp, color_rgb: int) -> None:
    """Remove 【】 brackets, then bold+color the bracketed keywords."""
    try:
        tf = _com_get(shp, "TextFrame", None)
        if tf is None:
            return
        tr = tf.TextRange
        full_text = tr.Text

        keywords = list(dict.fromkeys(re.findall(r'【([^】]+)】', full_text)))
        if not keywords:
            return

        tr.Text = re.sub(r'[【】]', '', full_text)

        for kw in keywords:
            start = 1
            while start <= tr.Length:
                found = tr.Find(kw, start)
                if found is None:
                    break
                found.Font.Bold = True
                found.Font.Color = color_rgb
                start = found.Start + found.Length
    except Exception:
        pass  # coloring is cosmetic — never fail the build


def _replace_image(slide, shp, img_path: str) -> None:
    """Replace an image shape with a picture from img_path.

    Inserts at native size first, then scales to fit within the original
    slot bounds while preserving aspect ratio (no stretching), then centers.
    """
    try:
        slot_left   = float(_com_get(shp, "Left",   0))
        slot_top    = float(_com_get(shp, "Top",    0))
        slot_width  = float(_com_get(shp, "Width",  100))
        slot_height = float(_com_get(shp, "Height", 100))
        name = _safe_text(_com_get(shp, "Name", ""))
        shp.Delete()

        # Insert at native size (-1 = use image's own dimensions)
        new_shp = slide.Shapes.AddPicture(
            FileName=img_path,
            LinkToFile=False,
            SaveWithDocument=True,
            Left=slot_left, Top=slot_top,
            Width=-1, Height=-1,
        )

        # Scale uniformly to fit within slot (no stretching)
        nat_w = float(_com_get(new_shp, "Width",  slot_width))
        nat_h = float(_com_get(new_shp, "Height", slot_height))
        scale = min(slot_width / nat_w, slot_height / nat_h)
        new_w = nat_w * scale
        new_h = nat_h * scale

        # Center within original slot
        new_shp.Width  = new_w
        new_shp.Height = new_h
        new_shp.Left   = slot_left  + (slot_width  - new_w) / 2
        new_shp.Top    = slot_top   + (slot_height - new_h) / 2

        try:
            new_shp.Name = name
        except Exception:
            pass
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def make_codex_slide(mc_sht, mc_ppt, mc_slide, sample_name: str,
                     mc_gpt: str = "n", mc_model: str = _MODEL):
    """Generate Codex evaluation slide, integrated into main.py Copy->Paste flow.

    Clones Slide(15) (standard template page) to end of presentation,
    then writes all shape content per CODEX_SHAPES specs.

    Returns the new slide object (caller should update mc_slide).
    """
    gpt_enabled = (mc_gpt == "y")
    rows = _xlwings_to_rows(mc_sht)

    # === Clone pattern — identical to all other sections in main.py ===
    X = mc_ppt.Slides.Count + 1
    mc_ppt.Slides(15).Copy()
    time.sleep(_COPY_PASTE_DELAY)
    new_slide = mc_ppt.Slides.Paste(X)
    time.sleep(1.0)

    # === Per-shape content build and write ===
    for spec in CODEX_SHAPES:
        name     = spec["name"]
        strategy = spec["strategy"]
        color_h  = spec.get("color_hint", "")

        if strategy == "skip":
            continue

        # Find shape on the new slide
        shp = None
        try:
            shp = new_slide.Shapes(name)
        except Exception:
            pass
        if shp is None:
            continue

        # Special case: image extraction needs the sheet reference
        if strategy == "extract_image":
            img_path = _extract_shoe_image(mc_sht)
            if img_path:
                _replace_image(new_slide, shp, img_path)
            continue

        # Build content
        content = _build_content(spec, rows, gpt_enabled, mc_model)

        # Route to correct writer
        if strategy == "mean_extraction" or bool(_com_get(shp, "HasChart", False)):
            _write_chart(shp, content)
        else:
            ok = _write_text(shp, content)
            if ok and color_h:
                color_rgb = _RED if color_h == "red" else _BLUE if color_h == "blue" else None
                if color_rgb is not None:
                    _apply_keyword_color(shp, color_rgb)

    return new_slide
